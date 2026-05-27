import calendar
from datetime import date, timedelta

import openpyxl
from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db import transaction
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from core.models import Asistencia, Estudiante, Observacion

from core.asistencia_utils import calendario_asistencia_estudiante
from core.participacion_utils import resumen_participacion
from core.puntaje_utils import total_bruto_estudiante

from .forms import RegistroPadreForm, SolicitudNuevoEnlaceForm, VincularHijoForm
from .models import InvitacionPadre, Padre, UserProfile, UserRole
from .permissions import puede_gestionar_invitacion
from .roles import respuesta_si_no_modo_padre
from .utils import estudiantes_por_cedula_normalizada, normalizar_cedula


def _vincular_por_invitacion(request, inv: InvitacionPadre, cedula_input: str, *, nuevo_usuario: User | None):
    """Valida cédula, enlaza todos los estudiantes con esa cédula + el del token; marca invitación usada."""
    norm = normalizar_cedula(cedula_input)
    invitado = inv.estudiante
    if normalizar_cedula(invitado.cedula) != norm:
        return False, "La cédula no coincide con la del estudiante de este enlace."
    matches = list(estudiantes_por_cedula_normalizada(norm))
    if invitado.pk not in {e.pk for e in matches}:
        matches.append(invitado)
    user = nuevo_usuario or request.user
    with transaction.atomic():
        if nuevo_usuario:
            padre, _ = Padre.objects.get_or_create(user=user)
            UserProfile.objects.update_or_create(
                user=user, defaults={"role": UserRole.PADRE}
            )
        else:
            padre = user.padre_profile
            UserProfile.objects.update_or_create(
                user=user, defaults={"role": UserRole.PADRE}
            )
        for e in matches:
            padre.estudiantes.add(e)
        inv.usado_en = timezone.now()
        inv.save(update_fields=["usado_en"])
    return True, None


@login_required
def padre_dashboard(request):
    redir = respuesta_si_no_modo_padre(request)
    if redir:
        return redir
    padre = request.user.padre_profile
    hijos = padre.estudiantes.all().order_by("grado", "apellido", "nombre")
    return render(request, "accounts/padre_dashboard.html", {"hijos": hijos})


@login_required
def padre_estudiante_curriculum(request, pk):
    redir = respuesta_si_no_modo_padre(request)
    if redir:
        return redir
    padre = request.user.padre_profile
    est = get_object_or_404(Estudiante, pk=pk)
    if est not in padre.estudiantes.all():
        raise Http404()
    hoy = date.today()
    try:
        mes = int(request.GET.get("mes", hoy.month))
        anio = int(request.GET.get("anio", hoy.year))
    except ValueError:
        mes, anio = hoy.month, hoy.year
    meses = [
        (1, "Enero"), (2, "Febrero"), (3, "Marzo"), (4, "Abril"),
        (5, "Mayo"), (6, "Junio"), (7, "Julio"), (8, "Agosto"),
        (9, "Septiembre"), (10, "Octubre"), (11, "Noviembre"), (12, "Diciembre"),
    ]
    asist_ctx = calendario_asistencia_estudiante(est, mes, anio)
    observaciones = (
        Observacion.objects.filter(estudiante=est, visible_padre=True)
        .select_related("materia")
        .order_by("-fecha")[:50]
    )
    try:
        from tareas.models import TareaEvaluacion
        from tareas.rubricas import promedio_normalizado

        evaluaciones = list(
            TareaEvaluacion.objects.filter(estudiante=est)
            .select_related("tarea")
            .order_by("-tarea__fecha_entrega", "-tarea__creado_en")[:100]
        )
        eval_con_puntaje = [e for e in evaluaciones if e.puntaje is not None]
        prom_tareas_pct = promedio_normalizado(eval_con_puntaje)
    except Exception:
        evaluaciones = []
        prom_tareas_pct = None

    participaciones = est.participaciones.select_related("materia").order_by("-fecha")
    pts = resumen_participacion(participaciones)
    score_bruto = total_bruto_estudiante(pts, prom_tareas_pct)

    return render(
        request,
        "accounts/padre_curriculum.html",
        {
            "estudiante": est,
            "meses": meses,
            "observaciones": observaciones,
            "evaluaciones": evaluaciones,
            "total_positivos": pts["total_positivos"],
            "total_negativos": pts["total_negativos"],
            "total_bonos": pts["total_bonos"],
            "balance_total": pts["balance_total"],
            "prom_tareas_pct": prom_tareas_pct,
            "score_bruto": score_bruto,
            **asist_ctx,
        },
    )


@login_required
def padre_exportar_asistencias(request, pk):
    redir = respuesta_si_no_modo_padre(request)
    if redir:
        return redir
    padre = request.user.padre_profile
    est = get_object_or_404(Estudiante, pk=pk)
    if est not in padre.estudiantes.all():
        raise Http404()
    hoy = date.today()
    try:
        mes = int(request.GET.get("mes", hoy.month))
        anio = int(request.GET.get("anio", hoy.year))
    except ValueError:
        mes, anio = hoy.month, hoy.year
    meses_nombres = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
        5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
        9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
    }
    fecha_inicio = date(anio, mes, 1)
    fecha_fin = date(anio, mes, calendar.monthrange(anio, mes)[1])
    dias_lectivos = []
    d = fecha_inicio
    while d <= fecha_fin:
        if d.weekday() < 5:
            dias_lectivos.append(d)
        d += timedelta(days=1)
    asist_map = {}
    for a in Asistencia.objects.filter(
        estudiante=est, fecha__gte=fecha_inicio, fecha__lte=fecha_fin
    ):
        asist_map[a.fecha] = a.estado
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencia"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    t = ws.cell(
        row=1,
        column=1,
        value=f"Asistencia — {est.apellido}, {est.nombre} — {meses_nombres[mes]} {anio}",
    )
    t.font = Font(bold=True, size=14)
    headers = ["Día lectivo", "Estado"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin
    for i, dia in enumerate(dias_lectivos, 4):
        ws.cell(row=i, column=1, value=dia.isoformat()).border = thin
        ws.cell(row=i, column=2, value=asist_map.get(dia, "-")).border = thin
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="Asistencia_{est.pk}_{mes}_{anio}.xlsx"'
    )
    wb.save(response)
    return response


@require_http_methods(["GET", "POST"])
def registro_padre(request, token):
    inv = get_object_or_404(InvitacionPadre, token=token)
    if inv.usado_en:
        return render(request, "accounts/invitacion_usada.html", {"inv": inv})
    if not inv.activa():
        form_sol = SolicitudNuevoEnlaceForm()
        if request.method == "POST" and "solicitud" in request.POST:
            form_sol = SolicitudNuevoEnlaceForm(request.POST)
            if form_sol.is_valid():
                form_sol.save()
                messages.success(
                    request,
                    "Recibimos su solicitud. El colegio le generará un nuevo enlace.",
                )
                return redirect("account_login")
        return render(
            request,
            "accounts/invitacion_expirada.html",
            {"inv": inv, "form_sol": form_sol},
        )
    form = RegistroPadreForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        email = form.cleaned_data["email"]
        try:
            with transaction.atomic():
                user = User.objects.create_user(
                    username=email,
                    email=email,
                    password=form.cleaned_data["password1"],
                )
                ok, err = _vincular_por_invitacion(
                    request, inv, form.cleaned_data["cedula"], nuevo_usuario=user
                )
                if not ok:
                    raise ValueError(err)
        except ValueError as e:
            messages.error(request, str(e))
        else:
            login(request, user, backend="django.contrib.auth.backends.ModelBackend")
            messages.success(request, "Cuenta creada y estudiante vinculado.")
            return redirect("accounts:padre_dashboard")
    return render(
        request,
        "accounts/registro_padre.html",
        {"form": form, "inv": inv, "estudiante": inv.estudiante},
    )


@login_required
@require_http_methods(["GET", "POST"])
def vincular_otro_hijo(request, token):
    inv = get_object_or_404(InvitacionPadre, token=token)
    redir = respuesta_si_no_modo_padre(request)
    if redir:
        return redir
    padre = request.user.padre_profile
    if inv.usado_en:
        return render(request, "accounts/invitacion_usada.html", {"inv": inv})
    if not inv.activa():
        form_sol = SolicitudNuevoEnlaceForm()
        if request.method == "POST" and "solicitud" in request.POST:
            form_sol = SolicitudNuevoEnlaceForm(request.POST)
            if form_sol.is_valid():
                form_sol.save()
                messages.success(request, "Solicitud enviada. Contacte al colegio si tiene prisa.")
                return redirect("accounts:padre_dashboard")
        return render(
            request,
            "accounts/invitacion_expirada.html",
            {"inv": inv, "form_sol": form_sol},
        )
    form = VincularHijoForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        ok, err = _vincular_por_invitacion(
            request, inv, form.cleaned_data["cedula"], nuevo_usuario=None
        )
        if not ok:
            messages.error(request, err)
        else:
            messages.success(request, "Estudiante vinculado correctamente.")
            return redirect("accounts:padre_dashboard")
    return render(
        request,
        "accounts/vincular_hijo.html",
        {"form": form, "inv": inv, "estudiante": inv.estudiante},
    )


@login_required
@require_http_methods(["POST"])
def crear_invitacion_estudiante(request, pk):
    est = get_object_or_404(Estudiante, pk=pk)
    if not puede_gestionar_invitacion(request.user, est):
        raise Http404()
    InvitacionPadre.crear(est, usuario=request.user)
    messages.success(request, "Se generó un nuevo enlace de invitación (válido 24 horas).")
    return redirect("core:estudiante_detail", pk=pk)


def privacidad_resumen(request):
    return render(request, "accounts/privacidad.html")

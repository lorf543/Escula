from datetime import date, timedelta
import calendar

from django.contrib.auth.decorators import login_required, user_passes_test
from django.conf import settings
from django.db.models import Count, Q
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST

from accounts.models import InvitacionPadre
from accounts.permissions import staff_o_profesor

from .forms import (
    AsistenciaForm,
    EstudianteForm,
    MateriaForm,
    ObservacionForm,
    ParticipacionForm,
    ProfesorForm,
)
from .models import (
    Asistencia, Estudiante, Materia, Observacion, Participacion, Profesor, GRADO_CHOICES, ESTADO_ASISTENCIA,
)
from .asistencia_utils import calendario_asistencia_estudiante
from .participacion_utils import resumen_participacion
from .puntaje_utils import total_bruto_estudiante


def school_access(view_fn):
    return login_required(
        user_passes_test(staff_o_profesor, login_url=settings.LOGIN_URL)(view_fn)
    )


# ── Dashboard ──────────────────────────────────────────────────────────


@school_access
def dashboard(request):
    ranking = []
    promedio_bruto = 0
    alertas_bajas = 0
    try:
        from tareas.models import TareaEvaluacion
        from tareas.rubricas import promedio_normalizado

        estudiantes = list(Estudiante.objects.all())
        total_sum = 0
        for est in estudiantes:
            part = resumen_participacion(est.participaciones.all())
            evals = list(
                TareaEvaluacion.objects.filter(estudiante=est, puntaje__isnull=False).select_related("tarea")
            )
            prom = promedio_normalizado(evals) if evals else None
            bruto = total_bruto_estudiante(part, prom)
            total_sum += bruto["total_bruto"]
            if bruto["total_bruto"] < 0:
                alertas_bajas += 1
            ranking.append({"estudiante": est, "total_bruto": bruto["total_bruto"]})
        ranking.sort(key=lambda x: x["total_bruto"], reverse=True)
        ranking = ranking[:5]
        promedio_bruto = round(total_sum / len(estudiantes), 1) if estudiantes else 0
    except Exception:
        ranking = []
        promedio_bruto = 0
        alertas_bajas = 0

    ctx = {
        "total_estudiantes": Estudiante.objects.count(),
        "total_profesores": Profesor.objects.count(),
        "total_materias": Materia.objects.count(),
        "total_observaciones": Observacion.objects.count(),
        "ultimas_observaciones": Observacion.objects.select_related("estudiante")[:5],
        "estudiantes_por_grado": (
            Estudiante.objects.values("grado")
            .annotate(total=Count("id"))
            .order_by("grado")
        ),
        "ranking_bruto": ranking,
        "promedio_bruto_global": promedio_bruto,
        "alertas_bruto_bajo": alertas_bajas,
    }
    return render(request, "core/dashboard.html", ctx)


# ── Materias CRUD ──────────────────────────────────────────────────────


@school_access
def materia_list(request):
    q = request.GET.get("q", "")
    materias = Materia.objects.all()
    if q:
        materias = materias.filter(
            Q(nombre__icontains=q) | Q(codigo__icontains=q)
        )
    if request.htmx:
        return render(request, "core/materias/table.html", {"materias": materias, "q": q})
    return render(request, "core/materias/list.html", {"materias": materias, "q": q})


@school_access
def materia_create(request):
    form = MateriaForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        return HttpResponse(status=204, headers={"HX-Trigger": "materiaChanged"})
    return render(request, "core/materias/form.html", {"form": form, "title": "Nueva Materia"})


@school_access
def materia_update(request, pk):
    materia = get_object_or_404(Materia, pk=pk)
    form = MateriaForm(request.POST or None, instance=materia)
    if request.method == "POST" and form.is_valid():
        form.save()
        return HttpResponse(status=204, headers={"HX-Trigger": "materiaChanged"})
    return render(request, "core/materias/form.html", {"form": form, "title": "Editar Materia"})


@school_access
def materia_delete(request, pk):
    materia = get_object_or_404(Materia, pk=pk)
    if request.method == "POST":
        materia.delete()
        return HttpResponse(status=204, headers={"HX-Trigger": "materiaChanged"})
    return render(request, "core/materias/delete.html", {"object": materia})


# ── Profesores CRUD ────────────────────────────────────────────────────


@school_access
def profesor_list(request):
    q = request.GET.get("q", "")
    profesores = Profesor.objects.prefetch_related("materias").all()
    if q:
        profesores = profesores.filter(
            Q(nombre__icontains=q) | Q(apellido__icontains=q) | Q(cedula__icontains=q)
        )
    if request.htmx:
        return render(request, "core/profesores/table.html", {"profesores": profesores, "q": q})
    return render(request, "core/profesores/list.html", {"profesores": profesores, "q": q})


@school_access
def profesor_create(request):
    form = ProfesorForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        return HttpResponse(status=204, headers={"HX-Trigger": "profesorChanged"})
    return render(request, "core/profesores/form.html", {"form": form, "title": "Nuevo Profesor"})


@school_access
def profesor_update(request, pk):
    profesor = get_object_or_404(Profesor, pk=pk)
    form = ProfesorForm(request.POST or None, instance=profesor)
    if request.method == "POST" and form.is_valid():
        form.save()
        return HttpResponse(status=204, headers={"HX-Trigger": "profesorChanged"})
    return render(request, "core/profesores/form.html", {"form": form, "title": "Editar Profesor"})


@school_access
def profesor_detail(request, pk):
    profesor = get_object_or_404(Profesor, pk=pk)
    return render(request, "core/profesores/detail.html", {"profesor": profesor})


@school_access
def profesor_delete(request, pk):
    profesor = get_object_or_404(Profesor, pk=pk)
    if request.method == "POST":
        profesor.delete()
        return HttpResponse(status=204, headers={"HX-Trigger": "profesorChanged"})
    return render(request, "core/profesores/delete.html", {"object": profesor})


# ── Estudiantes CRUD ───────────────────────────────────────────────────


@school_access
def estudiante_list(request):
    q = request.GET.get("q", "")
    grado = request.GET.get("grado", "")
    estudiantes = Estudiante.objects.prefetch_related("profesores").all()
    if q:
        estudiantes = estudiantes.filter(
            Q(nombre__icontains=q) | Q(apellido__icontains=q) | Q(cedula__icontains=q)
        )
    if grado:
        estudiantes = estudiantes.filter(grado=grado)
    ctx = {"estudiantes": estudiantes, "q": q, "grado": grado}
    if request.htmx:
        return render(request, "core/estudiantes/table.html", ctx)
    return render(request, "core/estudiantes/list.html", ctx)


@school_access
def estudiante_create(request):
    form = EstudianteForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        return HttpResponse(status=204, headers={"HX-Trigger": "estudianteChanged"})
    return render(
        request, "core/estudiantes/form.html", {"form": form, "title": "Nuevo Estudiante"}
    )


@school_access
def estudiante_update(request, pk):
    est = get_object_or_404(Estudiante, pk=pk)
    form = EstudianteForm(request.POST or None, instance=est)
    if request.method == "POST" and form.is_valid():
        form.save()
        return HttpResponse(status=204, headers={"HX-Trigger": "estudianteChanged"})
    return render(
        request, "core/estudiantes/form.html", {"form": form, "title": "Editar Estudiante"}
    )


def _get_materias_estudiante(est):
    """Get all materias available to a student through their assigned professors."""
    return Materia.objects.filter(profesores__in=est.profesores.all()).distinct()


def materias_curriculum_estudiante(est):
    """Materias del currículum: profesores asignados + materias con actividad registrada."""
    por_profesor = Materia.objects.filter(profesores__in=est.profesores.all())
    por_actividad = Materia.objects.filter(
        Q(participaciones__estudiante=est) | Q(observaciones__estudiante=est)
    )
    return (por_profesor | por_actividad).distinct().order_by("nombre")


def _profesor_materia_estudiante(est, materia):
    """Profesor del estudiante que imparte la materia, si existe."""
    if not materia:
        return None
    return est.profesores.filter(materias=materia).first()


@school_access
def estudiante_detail(request, pk):
    est = get_object_or_404(Estudiante, pk=pk)
    observaciones = est.observaciones.select_related("materia").all()
    participaciones = est.participaciones.select_related("materia").all()
    pts = resumen_participacion(participaciones)

    now = timezone.now()
    inv_activa = (
        InvitacionPadre.objects.filter(
            estudiante=est, usado_en__isnull=True, expira_en__gt=now
        )
        .order_by("-creado_en")
        .first()
    )
    url_invitacion = None
    if inv_activa:
        url_invitacion = request.build_absolute_uri(
            reverse("accounts:registro_padre", args=[inv_activa.token])
        )

    try:
        from tareas.models import TareaEvaluacion

        evaluaciones_tareas = (
            TareaEvaluacion.objects.filter(estudiante=est)
            .select_related("tarea")
            .order_by("-tarea__fecha_entrega")[:30]
        )
    except Exception:
        evaluaciones_tareas = []

    ctx = {
        "estudiante": est,
        "observaciones": observaciones,
        "participaciones": participaciones,
        "total_asistencias": est.asistencias.count(),
        "presentes": est.asistencias.filter(estado="P").count(),
        "pct_asistencia": round(
            est.asistencias.filter(estado="P").count() / est.asistencias.count() * 100, 1
        ) if est.asistencias.count() else 0,
        "puntos_por_materia": pts["puntos_por_materia"],
        "total_positivos": pts["total_positivos"],
        "total_negativos": pts["total_negativos"],
        "balance_total": pts["balance_total"],
        "url_invitacion_padre": url_invitacion,
        "invitacion_padre": inv_activa,
        "evaluaciones_tareas": evaluaciones_tareas,
    }
    return render(request, "core/estudiantes/detail.html", ctx)


def _usuario_puede_ver_curriculum_escuela(user, estudiante) -> bool:
    if not user.is_authenticated:
        return False
    if user.is_staff:
        return True
    pp = getattr(user, "profesor_perfil", None)
    if pp and estudiante in pp.profesor.estudiantes.all():
        return True
    return False


def _asistencias_ctx_mes(estudiante, mes, anio):
    _, last_day = calendar.monthrange(anio, mes)
    fecha_inicio = date(anio, mes, 1)
    fecha_fin = date(anio, mes, last_day)
    qs = (
        Asistencia.objects.filter(
            estudiante=estudiante,
            fecha__gte=fecha_inicio,
            fecha__lte=fecha_fin,
        )
        .order_by("fecha")
    )
    return {
        "mes": mes,
        "anio": anio,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "asistencias": qs,
    }


@school_access
def estudiante_curriculum(request, pk):
    """Resumen académico (asistencias del mes, participación, tareas, observaciones) para staff o profesor asignado."""
    est = get_object_or_404(Estudiante, pk=pk)
    if not _usuario_puede_ver_curriculum_escuela(request.user, est):
        raise Http404()
    hoy = date.today()
    try:
        mes = int(request.GET.get("mes", hoy.month))
        anio = int(request.GET.get("anio", hoy.year))
    except ValueError:
        mes, anio = hoy.month, hoy.year
    materia_param = request.GET.get("materia", "")
    materia_id = int(materia_param) if materia_param.isdigit() else None
    materias = list(materias_curriculum_estudiante(est))
    materia_sel = None
    if materia_id:
        materia_sel = next((m for m in materias if m.pk == materia_id), None)
        if materia_sel is None:
            materia_sel = get_object_or_404(Materia, pk=materia_id)
    profesor_materia = _profesor_materia_estudiante(est, materia_sel)

    meses = [
        (1, "Enero"), (2, "Febrero"), (3, "Marzo"), (4, "Abril"),
        (5, "Mayo"), (6, "Junio"), (7, "Julio"), (8, "Agosto"),
        (9, "Septiembre"), (10, "Octubre"), (11, "Noviembre"), (12, "Diciembre"),
    ]
    asist_ctx = calendario_asistencia_estudiante(est, mes, anio)

    observaciones_qs = est.observaciones.select_related("materia").order_by("-fecha")
    participaciones_qs = est.participaciones.select_related("materia").order_by("-fecha")
    if materia_id:
        observaciones_qs = observaciones_qs.filter(materia_id=materia_id)
        participaciones_qs = participaciones_qs.filter(materia_id=materia_id)

    observaciones = list(observaciones_qs[:5])
    pts = resumen_participacion(participaciones_qs)
    participaciones_recientes = list(participaciones_qs[:8])

    try:
        from tareas.models import ModalidadTarea, TareaEvaluacion
        from tareas.rubricas import promedio_normalizado

        eval_qs = (
            TareaEvaluacion.objects.filter(estudiante=est)
            .select_related("tarea", "tarea__materia", "tarea__creada_por")
            .order_by("-tarea__fecha_entrega", "-tarea__creado_en")
        )
        if materia_id:
            eval_qs = eval_qs.filter(
                Q(tarea__materia_id=materia_id)
                | Q(tarea__materia__isnull=True, tarea__creada_por__materias=materia_id)
            ).distinct()
        evaluaciones = list(eval_qs[:10])
        eval_con_puntaje = [e for e in evaluaciones if e.puntaje is not None]
        oral = [e for e in eval_con_puntaje if e.tarea.modalidad == ModalidadTarea.ORAL]
        esc = [e for e in eval_con_puntaje if e.tarea.modalidad == ModalidadTarea.ESCRITA]
        prom_oral = promedio_normalizado(oral)
        prom_escrita = promedio_normalizado(esc)
        prom_tareas_pct = promedio_normalizado(eval_con_puntaje)
    except Exception:
        evaluaciones = []
        prom_oral = prom_escrita = prom_tareas_pct = None

    score_bruto = total_bruto_estudiante(pts, prom_tareas_pct)
    part_con_bonos = pts["balance_total"] + pts["total_bonos"]

    return render(
        request,
        "core/estudiantes/curriculum.html",
        {
            "estudiante": est,
            "meses": meses,
            "materias": materias,
            "materia_id": materia_id,
            "materia_sel": materia_sel,
            "profesor_materia": profesor_materia,
            "observaciones": observaciones,
            "evaluaciones": evaluaciones,
            "participaciones_recientes": participaciones_recientes,
            "total_positivos": pts["total_positivos"],
            "total_negativos": pts["total_negativos"],
            "total_bonos": pts["total_bonos"],
            "balance_total": pts["balance_total"],
            "part_con_bonos": part_con_bonos,
            "puntos_por_materia": pts["puntos_por_materia"],
            "prom_oral": prom_oral,
            "prom_escrita": prom_escrita,
            "prom_tareas_pct": prom_tareas_pct,
            "score_bruto": score_bruto,
            **asist_ctx,
        },
    )


@school_access
def estudiante_delete(request, pk):
    est = get_object_or_404(Estudiante, pk=pk)
    if request.method == "POST":
        est.delete()
        return HttpResponse(status=204, headers={"HX-Trigger": "estudianteChanged"})
    return render(request, "core/estudiantes/delete.html", {"object": est})


# ── Observaciones ──────────────────────────────────────────────────────


@school_access
def observacion_create(request, estudiante_pk):
    est = get_object_or_404(Estudiante, pk=estudiante_pk)
    materias = _get_materias_estudiante(est)
    form = ObservacionForm(request.POST or None, materias_qs=materias)
    if request.method == "POST" and form.is_valid():
        obs = form.save(commit=False)
        obs.estudiante = est
        obs.save()
        return HttpResponse(status=204, headers={"HX-Trigger": "observacionChanged"})
    return render(
        request,
        "core/observaciones/form.html",
        {"form": form, "estudiante": est, "title": "Nueva Observación"},
    )


@school_access
def observacion_delete(request, pk):
    obs = get_object_or_404(Observacion, pk=pk)
    if request.method == "POST":
        obs.delete()
        return HttpResponse(status=204, headers={"HX-Trigger": "observacionChanged"})
    return render(request, "core/observaciones/delete.html", {"object": obs})


@school_access
def observacion_list_partial(request, estudiante_pk):
    est = get_object_or_404(Estudiante, pk=estudiante_pk)
    observaciones = est.observaciones.select_related("materia").all()
    return render(
        request,
        "core/observaciones/list_partial.html",
        {"observaciones": observaciones, "estudiante": est},
    )


# ── Participación ─────────────────────────────────────────────────────


@school_access
def participacion_create(request, estudiante_pk):
    est = get_object_or_404(Estudiante, pk=estudiante_pk)
    materias = _get_materias_estudiante(est)
    form = ParticipacionForm(request.POST or None, materias_qs=materias)
    if request.method == "POST" and form.is_valid():
        part = form.save(commit=False)
        part.estudiante = est
        part.save()
        return HttpResponse(status=204, headers={"HX-Trigger": "participacionChanged"})
    return render(
        request,
        "core/participaciones/form.html",
        {"form": form, "estudiante": est, "title": "Nuevo Punto"},
    )


@school_access
def participacion_delete(request, pk):
    part = get_object_or_404(Participacion, pk=pk)
    if request.method == "POST":
        part.delete()
        return HttpResponse(status=204, headers={"HX-Trigger": "participacionChanged"})
    return render(request, "core/participaciones/delete.html", {"object": part})


@school_access
def participacion_list_partial(request, estudiante_pk):
    est = get_object_or_404(Estudiante, pk=estudiante_pk)
    participaciones = est.participaciones.select_related("materia").all()
    pts = resumen_participacion(participaciones)

    return render(
        request,
        "core/participaciones/list_partial.html",
        {
            "participaciones": participaciones,
            "estudiante": est,
            "puntos_por_materia": pts["puntos_por_materia"],
            "total_bonos": pts["total_bonos"],
        },
    )


# ── Asistencia ─────────────────────────────────────────────────────────


@school_access
def asistencia_view(request):
    hoy = date.today()
    fecha_str = request.GET.get("fecha")
    grado = request.GET.get("grado", "")

    try:
        fecha = date.fromisoformat(fecha_str) if fecha_str else hoy
    except ValueError:
        fecha = hoy

    if not grado:
        ctx = {
            "fecha": fecha,
            "grado": grado,
            "grados": GRADO_CHOICES,
            "registros": [],
            "sin_grado": True,
        }
        return render(request, "core/asistencia/list.html", ctx)

    estudiantes = Estudiante.objects.filter(grado=grado).order_by("apellido", "nombre")

    existentes = {
        a.estudiante_id: a
        for a in Asistencia.objects.filter(
            estudiante__in=estudiantes, fecha=fecha
        )
    }

    ya_guardada = bool(existentes)

    registros = []
    for est in estudiantes:
        asist = existentes.get(est.pk)
        registros.append({"estudiante": est, "asistencia": asist})

    resumen = None
    if ya_guardada:
        estados = [r["asistencia"].estado for r in registros if r["asistencia"]]
        resumen = {
            "total": len(registros),
            "presentes": estados.count("P"),
            "ausentes": estados.count("A"),
            "tardanzas": estados.count("T"),
            "excusas": estados.count("E"),
        }

    ctx = {
        "fecha": fecha,
        "grado": grado,
        "grado_display": dict(GRADO_CHOICES).get(grado, grado),
        "registros": registros,
        "grados": GRADO_CHOICES,
        "sin_grado": False,
        "ya_guardada": ya_guardada,
        "resumen": resumen,
        "estados": ESTADO_ASISTENCIA,
    }

    if request.htmx:
        return render(request, "core/asistencia/table.html", ctx)
    return render(request, "core/asistencia/list.html", ctx)


@require_POST
@school_access
def asistencia_guardar(request):
    fecha_str = request.POST.get("fecha")
    grado = request.POST.get("grado", "")

    try:
        fecha = date.fromisoformat(fecha_str) if fecha_str else date.today()
    except ValueError:
        fecha = date.today()

    estudiantes = Estudiante.objects.filter(grado=grado)
    estados_validos = {v for v, _ in ESTADO_ASISTENCIA}

    for est in estudiantes:
        estado = request.POST.get(f"estado_{est.pk}", "P")
        if estado not in estados_validos:
            estado = "P"
        Asistencia.objects.update_or_create(
            estudiante=est,
            fecha=fecha,
            defaults={"estado": estado},
        )

    redirect_url = f"{request.build_absolute_uri('/')[:-1]}{reverse('core:asistencia')}?grado={grado}&fecha={fecha}"
    return redirect(redirect_url)


@require_POST
@school_access
def asistencia_toggle(request, pk):
    asist = get_object_or_404(Asistencia, pk=pk)
    estados = ["P", "A", "T", "E"]
    idx = estados.index(asist.estado) if asist.estado in estados else 0
    asist.estado = estados[(idx + 1) % len(estados)]
    asist.save()
    return render(request, "core/asistencia/badge.html", {"asistencia": asist})


@school_access
def asistencia_historial(request):
    grado = request.GET.get("grado", "")
    mes_str = request.GET.get("mes", "")
    anio_str = request.GET.get("anio", "")

    hoy = date.today()
    try:
        mes = int(mes_str) if mes_str else hoy.month
        anio = int(anio_str) if anio_str else hoy.year
    except ValueError:
        mes, anio = hoy.month, hoy.year

    meses = [
        (1, "Enero"), (2, "Febrero"), (3, "Marzo"), (4, "Abril"),
        (5, "Mayo"), (6, "Junio"), (7, "Julio"), (8, "Agosto"),
        (9, "Septiembre"), (10, "Octubre"), (11, "Noviembre"), (12, "Diciembre"),
    ]

    if not grado:
        ctx = {
            "grado": grado,
            "grados": GRADO_CHOICES,
            "mes": mes,
            "anio": anio,
            "meses": meses,
            "sin_grado": True,
        }
        return render(request, "core/asistencia/historial.html", ctx)

    _, last_day = calendar.monthrange(anio, mes)
    fecha_inicio = date(anio, mes, 1)
    fecha_fin = date(anio, mes, last_day)

    dias_lectivos = []
    d = fecha_inicio
    while d <= fecha_fin:
        if d.weekday() < 5:  # Lun-Vie
            dias_lectivos.append(d)
        d += timedelta(days=1)

    estudiantes = Estudiante.objects.filter(grado=grado).order_by("apellido", "nombre")

    asistencias_qs = Asistencia.objects.filter(
        estudiante__grado=grado,
        fecha__gte=fecha_inicio,
        fecha__lte=fecha_fin,
    ).select_related("estudiante")

    asist_map = {}
    for a in asistencias_qs:
        asist_map[(a.estudiante_id, a.fecha)] = a.estado

    filas = []
    for est in estudiantes:
        estados = []
        p_count = a_count = t_count = e_count = 0
        for dia in dias_lectivos:
            estado = asist_map.get((est.id, dia), "")
            estados.append(estado)
            if estado == "P":
                p_count += 1
            elif estado == "A":
                a_count += 1
            elif estado == "T":
                t_count += 1
            elif estado == "E":
                e_count += 1
        total_reg = p_count + a_count + t_count + e_count
        pct = round(p_count / total_reg * 100, 1) if total_reg else 0
        filas.append({
            "estudiante": est,
            "estados": estados,
            "presentes": p_count,
            "ausentes": a_count,
            "tardanzas": t_count,
            "excusas": e_count,
            "pct": pct,
        })

    ctx = {
        "grado": grado,
        "grado_display": dict(GRADO_CHOICES).get(grado, grado),
        "grados": GRADO_CHOICES,
        "mes": mes,
        "anio": anio,
        "meses": meses,
        "mes_display": dict(meses).get(mes, ""),
        "dias_lectivos": dias_lectivos,
        "filas": filas,
        "sin_grado": False,
        "total_estudiantes": len(filas),
    }

    if request.htmx:
        return render(request, "core/asistencia/historial_table.html", ctx)
    return render(request, "core/asistencia/historial.html", ctx)


@school_access
def asistencia_exportar(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    grado = request.GET.get("grado", "")
    mes_str = request.GET.get("mes", "")
    anio_str = request.GET.get("anio", "")

    hoy = date.today()
    try:
        mes = int(mes_str) if mes_str else hoy.month
        anio = int(anio_str) if anio_str else hoy.year
    except ValueError:
        mes, anio = hoy.month, hoy.year

    meses_nombres = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
        5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
        9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
    }

    _, last_day = calendar.monthrange(anio, mes)
    fecha_inicio = date(anio, mes, 1)
    fecha_fin = date(anio, mes, last_day)

    dias_lectivos = []
    d = fecha_inicio
    while d <= fecha_fin:
        if d.weekday() < 5:
            dias_lectivos.append(d)
        d += timedelta(days=1)

    if grado:
        estudiantes = Estudiante.objects.filter(grado=grado).order_by("apellido", "nombre")
        grados_export = [(grado, dict(GRADO_CHOICES).get(grado, grado), estudiantes)]
    else:
        grados_export = []
        for g_val, g_label in GRADO_CHOICES:
            ests = Estudiante.objects.filter(grado=g_val).order_by("apellido", "nombre")
            if ests.exists():
                grados_export.append((g_val, g_label, ests))

    asist_qs = Asistencia.objects.filter(
        fecha__gte=fecha_inicio, fecha__lte=fecha_fin,
    ).select_related("estudiante")
    if grado:
        asist_qs = asist_qs.filter(estudiante__grado=grado)

    # Misma clave que el historial: código P/A/T/E (no get_estado_display + [:1])
    asist_map = {}
    for a in asist_qs:
        asist_map[(a.estudiante_id, a.fecha)] = a.estado

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    estado_fills = {
        "P": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "A": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "T": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "E": PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"),
    }
    estado_fonts = {
        "P": Font(color="006100", size=10),
        "A": Font(color="9C0006", size=10),
        "T": Font(color="9C6500", size=10),
        "E": Font(color="003366", size=10),
    }

    for g_val, g_label, estudiantes in grados_export:
        ws = wb.create_sheet(title=g_label[:31])

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + len(dias_lectivos))
        title = ws.cell(
            row=1, column=1,
            value=f"Asistencia - {g_label} - {meses_nombres[mes]} {anio}",
        )
        title.font = Font(bold=True, size=14, color="2F5496")
        title.alignment = center

        headers = ["#", "Apellido", "Nombre"]
        for dia in dias_lectivos:
            headers.append(dia.strftime("%d"))
        headers += ["P", "A", "T", "E", "%"]

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin

        for i, est in enumerate(estudiantes):
            row = 4 + i
            ws.cell(row=row, column=1, value=i + 1).border = thin
            ws.cell(row=row, column=1).alignment = center
            ws.cell(row=row, column=2, value=est.apellido).border = thin
            ws.cell(row=row, column=3, value=est.nombre).border = thin

            p = a = t = e = 0
            for j, dia in enumerate(dias_lectivos):
                col = 4 + j
                estado_code = asist_map.get((est.id, dia), "")
                cell = ws.cell(row=row, column=col, value=estado_code if estado_code else "")
                cell.alignment = center
                cell.border = thin
                if estado_code in estado_fills:
                    cell.fill = estado_fills[estado_code]
                    cell.font = estado_fonts[estado_code]
                if estado_code == "P":
                    p += 1
                elif estado_code == "A":
                    a += 1
                elif estado_code == "T":
                    t += 1
                elif estado_code == "E":
                    e += 1

            summary_col = 4 + len(dias_lectivos)
            for offset, val in enumerate([p, a, t, e]):
                cell = ws.cell(row=row, column=summary_col + offset, value=val)
                cell.alignment = center
                cell.border = thin

            total_reg = p + a + t + e
            pct = round(p / total_reg * 100, 1) if total_reg else 0
            pct_cell = ws.cell(row=row, column=summary_col + 4, value=pct / 100)
            pct_cell.number_format = "0.0%"
            pct_cell.alignment = center
            pct_cell.border = thin

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 20

    mes_nombre = meses_nombres.get(mes, str(mes))
    grado_str = grado if grado else "Todos"
    filename = f"Asistencia_{grado_str}_{mes_nombre}_{anio}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
